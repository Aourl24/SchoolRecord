from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse
from .models import Subject, Class, Record, StudentRecord, Student
from .report import Report

def export_report_excel(request):
    subject_id = request.GET.get('subject')
    class_name = request.GET.get('class')
    batch      = request.GET.get("batch")
    term       = request.GET.get("term")
    sort_order = request.GET.get('sort', 'asc')

    try:
        # ── Get report data (returns a dict, not a tuple) ──────────────────
        result = Report.generate_report(subject_id, class_name, batch, term, sort_order)

        if not result.get('success'):
            raise ValueError(result.get('error', 'Report generation failed'))

        total_report  = result.get('total_report')
        subject_model = result.get('subject')
        terms         = result.get('terms') or []
        term          = result.get('term')   # may be None for "All"

        if not total_report or len(total_report) < 2:
            raise ValueError("No data available for export")

        # First item is the header structure, rest are student rows
        header_data   = total_report[0]
        students_data = total_report[1:]

        # ── Build workbook ──────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Report"

        bold_font        = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        wrap_alignment   = Alignment(wrap_text=True, horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        # ── Calculate column structure ──────────────────────────────────────
        col_position = 3  # columns 1=S/N, 2=Student, then terms start at 3

        term_columns = {}
        for term_name in terms:
            term_records  = header_data.get('term_headers', {}).get(term_name, [])
            record_count  = len(term_records)
            totals_count  = 2 if header_data.get('term_totals') else 0

            term_columns[term_name] = {
                'start_col':   col_position,
                'record_count': record_count,
                'totals_count': totals_count,
                'total_cols':   record_count + totals_count,
            }
            col_position += record_count + totals_count

        # +2 for Total Score and Percentage at the end
        total_columns = col_position + 1

        # ── Header Row 1: Term group headers ───────────────────────────────
        row = 1
        col = 1

        ws.cell(row=row, column=col, value="S/N")
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)
        col += 1

        ws.cell(row=row, column=col, value="Student")
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)
        col += 1

        for term_name in terms:
            term_info = term_columns[term_name]
            if term_info['record_count'] > 0:
                ws.cell(row=row, column=col, value=term_name)
                span = term_info['record_count']
                if span > 1:
                    ws.merge_cells(start_row=row, start_column=col,
                                   end_row=row, end_column=col + span - 1)
                col += span

            if header_data.get('term_totals') and term_info['totals_count'] > 0:
                ws.cell(row=row, column=col, value=f"{term_name} Totals")
                if term_info['totals_count'] > 1:
                    ws.merge_cells(start_row=row, start_column=col,
                                   end_row=row, end_column=col + term_info['totals_count'] - 1)
                col += term_info['totals_count']

        ws.cell(row=row, column=col, value="Total Score")
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)
        col += 1

        ws.cell(row=row, column=col, value="Percentage")
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)

        # ── Header Row 2: Individual assessment sub-headers ─────────────────
        row = 2
        col = 3

        for term_name in terms:
            term_records = header_data.get('term_headers', {}).get(term_name, [])
            for rec in term_records:
                ws.cell(row=row, column=col, value=f"{rec['type']} {rec['number']}")
                col += 1

            if header_data.get('term_totals'):
                ws.cell(row=row, column=col, value="Test Total");  col += 1
                ws.cell(row=row, column=col, value="Term Total");  col += 1

        # ── Style both header rows ──────────────────────────────────────────
        for row_num in [1, 2]:
            for col_num in range(1, total_columns + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font      = bold_font
                cell.alignment = wrap_alignment
                cell.border    = border

        # ── Data rows ──────────────────────────────────────────────────────
        for idx, student in enumerate(students_data, start=1):
            row += 1
            col  = 1

            # S/N
            cell = ws.cell(row=row, column=col, value=idx)
            cell.alignment = center_alignment
            cell.border    = border
            col += 1

            # Student name + batch
            cell = ws.cell(row=row, column=col,
                           value=f"{student.get('name','')} ({student.get('class_name','')})")
            cell.border = border
            col += 1

            # Scores per term
            for term_name in terms:
                term_records = student.get('record_by_term', {}).get(term_name, [])
                for rec in term_records:
                    score_value = rec['score'] if rec.get('score') not in ('-', None) else None
                    cell = ws.cell(row=row, column=col, value=score_value)
                    cell.alignment = center_alignment
                    cell.border    = border
                    col += 1

                if header_data.get('term_totals') and 'term_totals' in student:
                    term_data = student['term_totals'].get(term_name, {})
                    cell = ws.cell(row=row, column=col, value=term_data.get('test_total', 0))
                    cell.alignment = center_alignment; cell.border = border; col += 1

                    cell = ws.cell(row=row, column=col, value=term_data.get('total_score', 0))
                    cell.alignment = center_alignment; cell.border = border; col += 1

            # Total score
            cell = ws.cell(row=row, column=col, value=student.get('total_score', 0))
            cell.alignment = center_alignment; cell.border = border; col += 1

            # Percentage
            cell = ws.cell(row=row, column=col, value=f"{student.get('percentage', 0)}%")
            cell.alignment = center_alignment; cell.border = border

        # ── Column widths ───────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        for col_num in range(3, total_columns + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 12

        # ── Freeze S/N + Student columns and both header rows ───────────────
        ws.freeze_panes = 'C3'

        # ── Build response ──────────────────────────────────────────────────
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        batch_text   = batch if batch and batch != "All" else "All_Batches"
        term_text    = term  if term  else "All_Terms"
        subject_name = subject_model.name if subject_model else "Report"
        filename     = f"{class_name}_{batch_text}_{subject_name}_{term_text}_Report.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Error generating Excel report: {str(e)}", status=500)
